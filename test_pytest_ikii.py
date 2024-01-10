from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c 


class Test:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get("https://www.saucedemo.com")
        self.driver.maximize_window() 

    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()
    
    
    def getData(): #veri çağırdığımızda self operatörünü çağırmıyoruz.
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağını gösteriyoruz
        rows = sheet.max_row #kaçıncı satıra kadar veri var ona bakılır
        data = []

        for i in range(2,rows+1):#ilki username ve password yani başlık olduğu için wxcelin ikinci satırından başladık ve row a kadar aldık
            username = sheet.cell(i,1).value # username başlığın değerine ulaştık. ama içindeki değere de ulaşmak için value filtresi koyduk
            password = sheet.cell(i,2).value
            data.append((username,password))#username ve password bir data olduğu için tekrar parantez içine aldık
            
        return data 

    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH
        sleep(5)

    @pytest.mark.parametrize("username,password",[("1","secret_sauce"),("problem_user","1"),("error_user","1")])
    def test_invalid_login(self,username,password):
        
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)
        
        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()
        
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert  errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH
    
    @pytest.mark.skip("username,password",[("standard_user","secret_sauce")])
    def test_valid_remove(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()

        self.driver.execute_script("window.scrollTo(0,500)") 
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='add-to-cart-sauce-labs-fleece-jacket']")))
        addToCart.click()

        cartControl = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        cartControl.click()

        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='remove-sauce-labs-fleece-jacket']")))
        assert remove.text == "Remove"
        sleep(5)

    @pytest.mark.skip("username,password",[("standard_user","secret_sauce")])
    def test_filter(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)

        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)

        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()

        self.driver.execute_script("window.scrollTo(0,500)") 
        filterProduct = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[2]/div/span/select")))
        filterProduct.click()

        PriceLowToHigh = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[2]/div/span/select/option[3]")))
        PriceLowToHigh.click()
        sleep(5)

